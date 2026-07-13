"""
Google Sheet sync for the Standard 7 indicators tracker.

The team fills in a shared Google Sheet (one tab per standard, named
"Standard N" in any language — we only look for the trailing digit, so
"معيار 1".."معيار 7" works) with indicator status/responsible/evidence/due
date. The coordinator only needs the roll-up progress already served by
GET /api/indicators/summary; this module is what keeps indicators.py's
SQLite rows in sync with that sheet.

Access model: "anyone with the link can view" + the spreadsheet's public
XLSX export endpoint — no Google API credentials required. If the sheet
isn't shared that way, the export request comes back as an HTML sign-in
page instead of a workbook and we raise a clear error.

Matching is positional, not text-based: row N (in sheet order) within a
standard's tab maps to the Nth indicator (ordered by id) already stored
for that standard. This is deliberate — the sheet holds the real official
indicator wording while the tracker was seeded with placeholder text, so
matching by text would never line up. A sync also overwrites indicator_text
with the sheet's wording, which is exactly how the placeholders get
replaced with the real thing.
"""
from __future__ import annotations

import datetime
import io
import logging
import re
import urllib.error
import urllib.request
from typing import Any

import indicators

logger = logging.getLogger(__name__)

_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
_TAB_STANDARD_RE = re.compile(r"(\d+)")

_STATUS_MAP = {
    # Arabic — "not available / hasn't started"
    "غير متوفر": "missing",
    "غير مكتمل": "missing",
    "لم يبدأ": "missing",
    # Arabic — "in progress"
    "قيد التنفيذ": "partial",
    "تحت التنفيذ": "partial",
    "جزئي": "partial",
    "جزئية": "partial",
    # Arabic — "done"
    "مكتمل": "complete",
    "مكتملة": "complete",
    "منجز": "complete",
    "تم": "complete",
    # English, in case a team member types it directly
    "missing": "missing",
    "not available": "missing",
    "n/a": "missing",
    "partial": "partial",
    "complete": "complete",
    "done": "complete",
    "in progress": "partial",
}


def _now_iso() -> str:
    return datetime.datetime.now().isoformat()


def extract_sheet_id(sheet_url_or_id: str) -> str:
    """Accept either a full Google Sheets URL or a bare sheet ID."""
    m = _SHEET_ID_RE.search(sheet_url_or_id)
    return m.group(1) if m else sheet_url_or_id.strip()


def _fetch_workbook_bytes(sheet_id: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            content_type = resp.headers.get("Content-Type", "")
            content = resp.read()
    except urllib.error.URLError as exc:
        raise ValueError(f"Could not reach the Google Sheet: {exc}") from exc

    if "spreadsheet" not in content_type and "octet-stream" not in content_type:
        raise ValueError(
            "The sheet did not return a spreadsheet — make sure it's shared as "
            '"Anyone with the link can view".'
        )
    return content


def _standard_number_from_tab(tab_name: str) -> int | None:
    m = _TAB_STANDARD_RE.search(tab_name)
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 7 else None


def _flatten_row(row_values: tuple) -> list[str]:
    """Handle both a well-formed row (one value per real column) and this
    sheet's current format (everything tab-separated inside column A)."""
    parts: list[str] = []
    for v in row_values:
        if v is None:
            continue
        parts.extend(str(v).split("\t"))
    return [p.strip() for p in parts]


def _map_status(raw: str) -> str:
    key = raw.strip()
    if not key:
        return "missing"
    mapped = _STATUS_MAP.get(key) or _STATUS_MAP.get(key.lower())
    if mapped:
        return mapped
    logger.warning("Unrecognized status text from sheet: %r; defaulting to 'partial'", raw)
    return "partial"


_LEADING_INDEX_RE = re.compile(r"^(\d+)\s+(\S.*)$")


def _parse_sheet_row(parts: list[str]) -> dict[str, Any] | None:
    """parts: [index, indicator_text, status, responsible, evidence, due_date, last_updated, ...]

    Some tabs separate fields with a literal tab character (handled by
    _flatten_row already); others squash "<index>   <indicator text>" into
    a single cell with runs of spaces instead of a tab. Detect and split
    that case here before the normal positional parsing.
    """
    if not parts:
        return None
    if not parts[0].isdigit():
        m = _LEADING_INDEX_RE.match(parts[0])
        if not m:
            return None  # header or blank row
        parts = [m.group(1), m.group(2).strip()] + list(parts[1:])
    if not parts[0].isdigit():
        return None
    indicator_text = parts[1].strip() if len(parts) > 1 else ""
    if not indicator_text:
        return None
    return {
        "indicator_text": indicator_text,
        "status": _map_status(parts[2] if len(parts) > 2 else ""),
        "responsible_person": parts[3].strip() if len(parts) > 3 and parts[3].strip() else None,
        "evidence_link": parts[4].strip() if len(parts) > 4 and parts[4].strip() else None,
        "due_date": parts[5].strip() if len(parts) > 5 and parts[5].strip() else None,
    }


def sync_from_sheet(sheet_url_or_id: str) -> dict[str, Any]:
    """Fetch the sheet and sync every standard's tab into indicators.py.

    Returns a summary: {updated, added, skipped_tabs, warnings, synced_at}.
    """
    import openpyxl

    sheet_id = extract_sheet_id(sheet_url_or_id)
    content = _fetch_workbook_bytes(sheet_id)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    updated = 0
    added = 0
    skipped_tabs: list[str] = []
    warnings: list[str] = []

    for tab_name in wb.sheetnames:
        standard_number = _standard_number_from_tab(tab_name)
        if standard_number is None:
            skipped_tabs.append(tab_name)
            continue

        sheet_rows: list[dict[str, Any]] = []
        for row_idx, row in enumerate(wb[tab_name].iter_rows(values_only=True), start=1):
            parsed = _parse_sheet_row(_flatten_row(row))
            if parsed is not None:
                sheet_rows.append(parsed)

        existing = indicators.list_indicators(standard_number=standard_number)
        for i, parsed in enumerate(sheet_rows):
            if i < len(existing):
                indicators.update_indicator(existing[i]["id"], **parsed)
                updated += 1
            else:
                created = indicators.create_indicator(
                    standard_number=standard_number,
                    indicator_text=parsed["indicator_text"],
                    responsible_person=parsed["responsible_person"],
                    evidence_link=parsed["evidence_link"],
                    due_date=parsed["due_date"],
                )
                if parsed["status"] != "missing":
                    indicators.update_indicator(created["id"], status=parsed["status"])
                added += 1

    return {
        "updated": updated,
        "added": added,
        "skipped_tabs": skipped_tabs,
        "warnings": warnings,
        "synced_at": _now_iso(),
    }
