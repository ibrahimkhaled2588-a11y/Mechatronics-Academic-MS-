"""Smoke test for backend/sheets_sync.py.

Builds a synthetic in-memory workbook that mirrors the two row formats the
real team sheet actually uses (tab-separated fields in one cell for most
tabs, and "<index><spaces><text>" squashed into one cell for the "Standard
1" tab specifically) and monkeypatches the network fetch, so this test
never hits the real Google Sheet. Plain-script convention (see
test_indicators.py) — run directly: `python tests/test_sheets_sync.py`.
"""
import io
import os
import sys
import tempfile

backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(backend_dir)

_tmp_db = os.path.join(tempfile.gettempdir(), "test_sheets_sync.db")
if os.path.exists(_tmp_db):
    os.remove(_tmp_db)
os.environ["ACCREDITATION_DB_PATH"] = _tmp_db

import indicators  # noqa: E402
import sheets_sync  # noqa: E402
import openpyxl  # noqa: E402


def _build_fake_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Tab-separated format (matches Standard 2-7 in the real sheet)
    ws2 = wb.create_sheet("Standard 2")
    ws2["A1"] = "Indicator\tStatus\tResponsible\tEvidence\tDue date\tLast updated"
    ws2["A2"] = "1\tProgram specification matches national reference standards\tIn Progress\tDr. X\tDraft uploaded\t2026-06-01\t"
    ws2["A3"] = "2\tILOs are mapped to a course matrix\tNot Available\t\t\t\t"

    # Space-squashed format (matches Standard 1 in the real sheet)
    ws1 = wb.create_sheet("Standard 1")
    ws1["A1"] = "Indicator"
    ws1["A2"] = "1        Program mission is approved and documented                   "
    ws1["A3"] = "2        Organizational structure is documented                      "

    # A non-standard tab that should be skipped, not crash the sync
    wb.create_sheet("Notes")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_FAKE_BYTES = _build_fake_workbook_bytes()
sheets_sync._fetch_workbook_bytes = lambda sheet_id: _FAKE_BYTES  # monkeypatch: no network call

# --- row parsing unit checks ---
assert sheets_sync._standard_number_from_tab("Standard 2") == 2
assert sheets_sync._standard_number_from_tab("معيار 7") == 7
assert sheets_sync._standard_number_from_tab("Notes") is None

assert sheets_sync._map_status("Not Available") == "missing"
assert sheets_sync._map_status("غير متوفر") == "missing"
assert sheets_sync._map_status("In Progress") == "partial"
assert sheets_sync._map_status("قيد التنفيذ") == "partial"
assert sheets_sync._map_status("مكتمل") == "complete"
assert sheets_sync._map_status("") == "missing"
assert sheets_sync._map_status("some unrecognized text") == "partial"

parsed = sheets_sync._parse_sheet_row(["1        Program mission is approved and documented"])
assert parsed["indicator_text"] == "Program mission is approved and documented"
assert parsed["status"] == "missing"

assert sheets_sync._parse_sheet_row(["Indicator", "Status"]) is None  # header row
assert sheets_sync._parse_sheet_row([]) is None

# --- seed then sync ---
indicators.seed_defaults()
result = sheets_sync.sync_from_sheet("https://docs.google.com/spreadsheets/d/FAKE_ID/edit")
assert result["updated"] == 4, f"expected 4 rows updated (2 for Standard 1, 2 for Standard 2), got {result}"
assert result["added"] == 0
assert "Notes" in result["skipped_tabs"]

std1 = indicators.list_indicators(standard_number=1)
assert std1[0]["indicator_text"] == "Program mission is approved and documented"
assert std1[1]["indicator_text"] == "Organizational structure is documented"
assert std1[0]["status"] == "missing"
# Standard 1 has 4 real seeded indicators; the sheet only synced 2 rows, so
# the 3rd/4th must be untouched -- still the real official wording, not
# overwritten with something from the sheet.
assert std1[2]["indicator_text"] == indicators._SEED_INDICATORS[1][2]
assert std1[3]["indicator_text"] == indicators._SEED_INDICATORS[1][3]

std2 = indicators.list_indicators(standard_number=2)
assert std2[0]["indicator_text"] == "Program specification matches national reference standards"
assert std2[0]["status"] == "partial"
assert std2[0]["responsible_person"] == "Dr. X"
assert std2[0]["evidence_link"] == "Draft uploaded"
assert std2[0]["due_date"] == "2026-06-01"
assert std2[1]["status"] == "missing"
assert std2[1]["responsible_person"] is None, "blank sheet cell must not overwrite with an empty string"

# --- re-sync is idempotent (positional matching against the same existing rows) ---
result2 = sheets_sync.sync_from_sheet("https://docs.google.com/spreadsheets/d/FAKE_ID/edit")
assert result2["updated"] == 4
assert len(indicators.list_indicators(standard_number=1)) == 4, "re-sync must not create duplicate rows"
assert len(indicators.list_indicators(standard_number=2)) == 4

os.remove(_tmp_db)
print("All sheets_sync.py tests passed.")
