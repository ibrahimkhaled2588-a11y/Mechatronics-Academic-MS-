"""
FastAPI entry point for the Academic Data Quality & Analytics Service.
"""
from __future__ import annotations

import datetime
import io
import logging
import os
import sys
import uuid

from typing import List

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from pydantic import BaseModel
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

# Ensure backend directory is importable when run as a script
_base = os.path.dirname(__file__)
if _base not in sys.path:
    sys.path.insert(0, _base)

import pandas as pd

from quality import compute_kpis
from ml_models import course_risk_probabilities, predict_quality_degradation
from alerts import generate_alerts, generate_alerts_detailed
from academic_analytics import compute_academic_analytics
from course_plans import generate_course_plans
from statistics_program import compute_program_statistics
from statistics_cross import compute_cross_module_and_executive
from config import get_server, get_alert_thresholds
from chatbot import answer_question
import indicators
import curriculum_mapping
import governance

logger = logging.getLogger(__name__)

_server = get_server()
_thresholds = get_alert_thresholds()

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_frontend_dir = os.path.normpath(os.path.join(_base, "..", "frontend"))
if not os.path.isdir(_frontend_dir):
    _frontend_dir = os.path.join(_base, "frontend")

_template_path = os.path.normpath(
    os.path.join(_base, "..", "MECE Program Report Model.docx")
)

# ---------------------------------------------------------------------------
# Application
# ---------------------------------------------------------------------------
app = FastAPI(title="Predictive Data Precision & Academic Analytics Service")

# SECURITY: explicit origin list; credentials disabled so wildcard is not needed
app.add_middleware(
    CORSMiddleware,
    allow_origins=_server.allowed_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PATCH", "DELETE"],
    allow_headers=["Content-Type"],
)

app.mount("/static", StaticFiles(directory=_frontend_dir), name="static")


@app.on_event("startup")
def _seed_accreditation_defaults() -> None:
    indicators.seed_defaults()

# ---------------------------------------------------------------------------
# In-memory upload history  (bounded to prevent memory leak)
# ---------------------------------------------------------------------------
_upload_history: list[dict] = []


def _record_upload(meta: dict) -> None:
    _upload_history.append(meta)
    if len(_upload_history) > _server.max_upload_history:
        _upload_history.pop(0)


# ---------------------------------------------------------------------------
# File validation helpers
# ---------------------------------------------------------------------------
_ALLOWED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",  # some browsers send this for .xlsx
}


def _validate_file(f: UploadFile) -> None:
    name = f.filename or ""
    if not name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are accepted")


async def _read_limited(f: UploadFile) -> bytes:
    """Read upload with a hard size cap to prevent DoS."""
    max_bytes = _server.max_upload_mb * 1024 * 1024
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await f.read(65536)  # 64 KB chunks
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"File exceeds the {_server.max_upload_mb} MB size limit",
            )
        chunks.append(chunk)
    return b"".join(chunks)


# ---------------------------------------------------------------------------
# Page serving  (SECURITY: whitelist prevents path traversal)
# ---------------------------------------------------------------------------

@app.get("/")
def root():
    index_path = os.path.join(_frontend_dir, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return HTMLResponse(
        "<html><body><h1>Academic Analytics API</h1>"
        "<p><a href='/docs'>OpenAPI docs</a></p></body></html>"
    )


@app.get("/{page}.html")
def serve_html(page: str):
    page_name = f"{page}.html"
    if page_name not in _server.allowed_pages:
        raise HTTPException(status_code=404, detail="Page not found")
    return FileResponse(os.path.join(_frontend_dir, page_name))


@app.get("/api-info")
def api_info():
    return {"service": "Academic Analytics API", "docs": "/docs"}


# ---------------------------------------------------------------------------
# Excel reading helpers
# ---------------------------------------------------------------------------

def _read_excel_bytes(contents: bytes) -> dict[str, pd.DataFrame]:
    try:
        from excel_reader import read_excel_adaptive
        return read_excel_adaptive(io.BytesIO(contents))
    except Exception as exc:
        logger.warning("Adaptive reader failed, falling back to pandas: %s", exc)
        try:
            return pd.read_excel(io.BytesIO(contents), sheet_name=None, engine="openpyxl")
        except Exception as exc2:
            raise HTTPException(
                status_code=400, detail=f"Could not read Excel file: {exc2}"
            ) from exc2


def _extract_sheet_titles(contents: bytes) -> dict[str, dict]:
    """Best-effort academic year / semester / program banner extraction; never fatal."""
    try:
        from excel_reader import extract_sheet_title_metadata
        return extract_sheet_title_metadata(io.BytesIO(contents))
    except Exception as exc:
        logger.warning("Sheet title metadata extraction failed: %s", exc)
        return {}


# ---------------------------------------------------------------------------
# Core analytics pipeline
# ---------------------------------------------------------------------------

def _run_analytics(dfs: dict[str, pd.DataFrame], filename: str, sheet_titles: dict | None = None) -> dict:
    schema_info = {s: list(df.columns) for s, df in dfs.items()}
    kpi_results = {s: compute_kpis(df) for s, df in dfs.items()}
    academic = {s: compute_academic_analytics(df, {}) for s, df in dfs.items()}

    predictions = predict_quality_degradation(kpi_results)
    course_risk = course_risk_probabilities(academic)
    plans = generate_course_plans(academic, course_risk, predictions)
    program_stats = compute_program_statistics(academic, kpi_results)
    cross_stats = compute_cross_module_and_executive(kpi_results, predictions, None, program_stats)

    from ml_models import forecast_academic_trends
    trend_forecast = forecast_academic_trends(academic)

    _threshold_dict = {
        "duplicate_rate": _thresholds.duplicate_rate,
        "anomaly_density": _thresholds.anomaly_density,
        "quality_score": _thresholds.quality_score,
        "drift_score": _thresholds.drift_score,
        "predicted_risk_pct": _thresholds.predicted_risk_pct,
        "failure_rate": _thresholds.failure_rate,
    }
    alerts_detailed = generate_alerts_detailed(
        kpi_results, predictions, _threshold_dict,
        academic_analytics=academic, course_risk=course_risk,
    )
    alerts_list = [a["message"] for a in alerts_detailed]

    meta = {
        "filename": filename,
        "timestamp": pd.Timestamp.now().isoformat(),
        "sheets": list(dfs.keys()),
        "schema": schema_info,
        "kpis": kpi_results,
        "sheet_titles": sheet_titles or {},
    }
    _record_upload(meta)

    return jsonable_encoder({
        "metadata": meta,
        "predictions": predictions,
        "alerts": alerts_list,
        "alerts_detailed": alerts_detailed,
        "academic_analytics": academic,
        "course_risk": course_risk,
        "course_plans": plans,
        "program_statistics": program_stats,
        "cross_module": cross_stats,
        "trend_forecast": trend_forecast,
    })


# ---------------------------------------------------------------------------
# Upload endpoints
# ---------------------------------------------------------------------------

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Accept a single XLSX file and return full analytics."""
    _validate_file(file)
    contents = await _read_limited(file)
    try:
        dfs = _read_excel_bytes(contents)
        if not dfs:
            raise HTTPException(status_code=422, detail="No valid sheets found in file")
        sheet_titles = _extract_sheet_titles(contents)
        return JSONResponse(content=_run_analytics(dfs, file.filename or "upload.xlsx", sheet_titles))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error processing %s", file.filename)
        raise HTTPException(status_code=500, detail="Analysis failed; check server logs") from exc


def _merge_bylaw_dfs(sheets_old: dict, sheets_new: dict, prefix: str) -> dict:
    from merge_excel_stats import merge_sheet
    result = {}
    for name in sorted(set(sheets_old) | set(sheets_new)):
        merged = merge_sheet(
            sheets_old.get(name, pd.DataFrame()),
            sheets_new.get(name, pd.DataFrame()),
        )
        if not merged.empty:
            key = f"{prefix} | {name}" if prefix else name
            result[key] = merged
    return result


@app.post("/upload-excel-multi")
async def upload_excel_multi(
    sem1_old: UploadFile = File(...),
    sem1_new: UploadFile = File(...),
    sem2_old: UploadFile | None = File(None),
    sem2_new: UploadFile | None = File(None),
):
    """Accept 2 or 4 XLSX files (old/new bylaw per semester), merge, and return analytics."""
    for f in [sem1_old, sem1_new, sem2_old, sem2_new]:
        if f and f.filename:
            _validate_file(f)
    try:
        has_sem2 = sem2_old is not None and sem2_new is not None and sem2_old.filename and sem2_new.filename
        prefix1 = "Semester 1" if has_sem2 else ""

        contents1_new = await _read_limited(sem1_new)
        sheets1_old = _read_excel_bytes(await _read_limited(sem1_old))
        sheets1_new = _read_excel_bytes(contents1_new)
        dfs = _merge_bylaw_dfs(sheets1_old, sheets1_new, prefix1)
        filenames = [sem1_old.filename, sem1_new.filename]
        sheet_titles = _extract_sheet_titles(contents1_new)

        if has_sem2:
            contents2_new = await _read_limited(sem2_new)
            sheets2_old = _read_excel_bytes(await _read_limited(sem2_old))
            sheets2_new = _read_excel_bytes(contents2_new)
            dfs.update(_merge_bylaw_dfs(sheets2_old, sheets2_new, "Semester 2"))
            filenames += [sem2_old.filename, sem2_new.filename]
            for k, v in _extract_sheet_titles(contents2_new).items():
                sheet_titles.setdefault(k, v)

        return JSONResponse(content=_run_analytics(dfs, "; ".join(filenames), sheet_titles))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Multi-file upload failed")
        raise HTTPException(status_code=500, detail="Analysis failed; check server logs") from exc


@app.post("/upload-excel-bulk")
async def upload_excel_bulk(files: List[UploadFile] = File(...)):
    """Accept 2–10 XLSX files, merge their sheets by name, and return full analytics."""
    from merge_excel_stats import merge_sheet
    if not files or len(files) < 2:
        raise HTTPException(status_code=422, detail="Please upload at least 2 Excel files")
    if len(files) > 10:
        raise HTTPException(status_code=422, detail="Maximum 10 files allowed per upload")
    for f in files:
        _validate_file(f)
    try:
        merged: dict[str, pd.DataFrame] = {}
        filenames: list[str] = []
        sheet_titles: dict[str, dict] = {}
        for f in files:
            contents = await _read_limited(f)
            sheets = _read_excel_bytes(contents)
            filenames.append(f.filename or "upload.xlsx")
            for k, v in _extract_sheet_titles(contents).items():
                sheet_titles.setdefault(k, v)
            for sheet_name, df in sheets.items():
                if sheet_name in merged:
                    merged[sheet_name] = merge_sheet(merged[sheet_name], df)
                else:
                    merged[sheet_name] = df
        if not merged:
            raise HTTPException(status_code=422, detail="No valid sheets found in uploaded files")
        return JSONResponse(content=_run_analytics(merged, "; ".join(filenames), sheet_titles))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Bulk upload failed")
        raise HTTPException(status_code=500, detail="Analysis failed; check server logs") from exc


@app.post("/export-course-docx")
async def export_course_docx(request: Request):
    """Accept course-report indicators from the frontend and return a filled-in DOCX."""
    try:
        payload = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON body") from exc

    if not isinstance(payload, dict) or not payload.get("course"):
        raise HTTPException(status_code=422, detail="A 'course' name is required")

    try:
        from course_report_docx import build_course_report_docx
        docx_bytes = build_course_report_docx(payload)
    except Exception as exc:
        logger.exception("Course report DOCX generation failed")
        raise HTTPException(status_code=500, detail="Report generation failed; check server logs") from exc

    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in str(payload.get("course")))[:60].strip() or "course"
    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="Course_Report_{safe_name}.docx"'},
    )


@app.post("/export-course-docx-bulk")
async def export_course_docx_bulk(request: Request):
    """Accept a list of course-report payloads and return one DOCX per course, zipped."""
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON body") from exc

    courses = body.get("courses") if isinstance(body, dict) else None
    if not isinstance(courses, list) or not courses:
        raise HTTPException(status_code=422, detail="A non-empty 'courses' list is required")
    if len(courses) > 200:
        raise HTTPException(status_code=422, detail="Maximum 200 courses per bulk export")

    import zipfile
    from course_report_docx import build_course_report_docx

    buf = io.BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for payload in courses:
            if not isinstance(payload, dict) or not payload.get("course"):
                continue
            try:
                docx_bytes = build_course_report_docx(payload)
            except Exception:
                logger.exception("Skipping course in bulk export: %s", payload.get("course"))
                continue
            safe_name = "".join(
                c if c.isalnum() or c in " -_" else "_" for c in str(payload["course"])
            )[:60].strip() or "course"
            name = f"Course_Report_{safe_name}.docx"
            n = 2
            while name in used_names:
                name = f"Course_Report_{safe_name}_{n}.docx"
                n += 1
            used_names.add(name)
            zf.writestr(name, docx_bytes)

    if not used_names:
        raise HTTPException(status_code=422, detail="No valid courses could be processed")

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="Course_Reports.zip"'},
    )


@app.post("/export-program-docx")
async def export_program_docx(request: Request):
    """Accept frontend analysis JSON and return a filled-in DOCX report."""
    try:
        body = await request.json()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON body") from exc

    analysis = body.get("analysis") or {}
    manual_total = body.get("manual_total_students")
    if manual_total is not None:
        try:
            manual_total = float(manual_total)
        except (TypeError, ValueError):
            manual_total = None

    if not os.path.isfile(_template_path):
        logger.error("DOCX template not found at %s", _template_path)
        raise HTTPException(status_code=500, detail="Program report template not found on server")

    try:
        from program_report_docx import build_program_report_docx
        docx_bytes = build_program_report_docx(
            _template_path, analysis, manual_total_students=manual_total
        )
    except Exception as exc:
        logger.exception("DOCX generation failed")
        raise HTTPException(status_code=500, detail="Report generation failed; check server logs") from exc

    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": 'attachment; filename="Program_Report.docx"'},
    )


# ---------------------------------------------------------------------------
# Survey dashboard routes
# ---------------------------------------------------------------------------

_SURVEY_EXPORTS_DIR = os.path.join(_base, "exports", "survey_dashboards")
_GOVERNANCE_DOCS_DIR = os.path.join(_base, "exports", "governance_documents")


@app.post("/api/survey/dashboard")
async def api_survey_dashboard(files: List[UploadFile] = File(...)):
    """Accept survey .xlsx files and return a dashboard PNG + PPTX + ZIP package."""
    from survey_dashboard import (
        SurveyResult,
        build_dashboard,
        build_export_zip,
        calculate_satisfaction,
        export_ppt,
        generate_charts,
        load_data,
    )

    if not files:
        raise HTTPException(status_code=422, detail="No files uploaded")

    session_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
    out_dir = os.path.join(_SURVEY_EXPORTS_DIR, session_id)
    os.makedirs(out_dir, exist_ok=True)

    survey_results: list = []
    png_paths: list[str] = []

    for f in files:
        if not f.filename:
            continue
        contents = await f.read()
        survey_name = os.path.splitext(f.filename)[0]
        try:
            df = load_data(contents)
        except Exception as exc:
            logger.warning("Could not read survey file %s: %s", f.filename, exc)
            continue
        question_stats = generate_charts(df)
        overall_satisfaction = calculate_satisfaction(df)
        png_path = os.path.join(out_dir, f"{survey_name}_dashboard.png")
        try:
            insights, improvements = build_dashboard(
                survey_name, question_stats, overall_satisfaction, png_path
            )
        except Exception as exc:
            logger.warning("Dashboard build failed for %s: %s", survey_name, exc)
            insights, improvements = [], []
        survey_results.append(
            SurveyResult(
                survey_name=survey_name,
                question_stats=question_stats,
                overall_satisfaction=overall_satisfaction,
                dashboard_png_path=png_path,
                insights=insights,
                improvements=improvements,
            )
        )
        if os.path.isfile(png_path):
            png_paths.append(png_path)

    if not survey_results:
        raise HTTPException(status_code=422, detail="No valid survey files could be processed")

    pptx_path = os.path.join(out_dir, "survey_dashboard_report.pptx")
    try:
        export_ppt(survey_results, pptx_path)
    except Exception as exc:
        logger.exception("PPT export failed: %s", exc)
        raise HTTPException(status_code=500, detail="PowerPoint generation failed") from exc

    zip_path = os.path.join(out_dir, "survey_outputs.zip")
    build_export_zip([p for p in [pptx_path] + png_paths if os.path.isfile(p)], zip_path)

    base_url = f"/api/survey/exports/{session_id}"
    return JSONResponse(
        {
            "surveys_processed": len(survey_results),
            "zip_url": f"{base_url}/survey_outputs.zip",
            "pptx_url": f"{base_url}/survey_dashboard_report.pptx",
            "dashboard_images": [f"{base_url}/{os.path.basename(p)}" for p in png_paths],
        }
    )


@app.get("/api/survey/exports/{session_id}/{filename}")
async def api_survey_export(session_id: str, filename: str):
    """Serve a previously generated survey export file."""
    if ".." in session_id or "/" in session_id or ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Invalid path")
    path = os.path.join(_SURVEY_EXPORTS_DIR, session_id, filename)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="Export file not found")
    return FileResponse(path)


@app.get("/api/survey/template")
async def api_survey_template():
    """Return a blank survey template Excel file."""
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey"
        headers = [
            "Q1 جودة المقررات الدراسية",
            "Q2 كفاءة أعضاء هيئة التدريس",
            "Q3 خدمات المعامل والتجهيزات",
            "Q4 الخدمات الطلابية والإرشاد الأكاديمي",
            "Q5 مدى الاستفادة من البرنامج بشكل عام",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        responses = ["موافق تماماً", "موافق", "محايد", "غير موافق", "غير موافق تماماً"]
        # Two sample rows so the user sees the expected format
        for row_idx in range(2, 4):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx, value=responses[(row_idx + col_idx) % len(responses)])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="survey_template.xlsx"'},
        )
    except Exception as exc:
        logger.exception("Template generation failed")
        raise HTTPException(status_code=500, detail="Template generation failed") from exc


# ---------------------------------------------------------------------------
# Accreditation — Standard 7 Indicators Tracker
# ---------------------------------------------------------------------------
class IndicatorCreate(BaseModel):
    standard_number: int
    indicator_text: str
    responsible_person: str | None = None
    evidence_link: str | None = None
    due_date: str | None = None


class IndicatorUpdate(BaseModel):
    status: str | None = None
    responsible_person: str | None = None
    evidence_link: str | None = None
    due_date: str | None = None
    indicator_text: str | None = None


class LoopLogEntryCreate(BaseModel):
    weakness_identified: str
    action_taken: str | None = None
    entry_status: str | None = None
    entry_date: str | None = None


@app.get("/api/indicators/standards")
def api_indicators_standards():
    """Return the 7 standard numbers/names for building the tracker UI."""
    return [{"standard_number": n, "standard_name": name} for n, name in indicators.STANDARDS.items()]


@app.get("/api/indicators/summary")
def api_indicators_summary():
    """Per-standard status counts, for a progress overview."""
    return [s.__dict__ for s in indicators.summarize_by_standard()]


@app.get("/api/indicators")
def api_indicators_list(standard_number: int | None = None, status: str | None = None):
    if status is not None and status not in indicators.VALID_STATUSES:
        raise HTTPException(status_code=422, detail=f"status must be one of {indicators.VALID_STATUSES}")
    return indicators.list_indicators(standard_number=standard_number, status=status)


@app.get("/api/indicators/{indicator_id}")
def api_indicators_get(indicator_id: int):
    result = indicators.get_indicator(indicator_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Indicator not found")
    return result


@app.post("/api/indicators")
def api_indicators_create(body: IndicatorCreate):
    try:
        return indicators.create_indicator(
            standard_number=body.standard_number,
            indicator_text=body.indicator_text,
            responsible_person=body.responsible_person,
            evidence_link=body.evidence_link,
            due_date=body.due_date,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.patch("/api/indicators/{indicator_id}")
def api_indicators_update(indicator_id: int, body: IndicatorUpdate):
    try:
        result = indicators.update_indicator(indicator_id, **body.model_dump(exclude_unset=True))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if result is None:
        raise HTTPException(status_code=404, detail="Indicator not found")
    return result


@app.post("/api/indicators/{indicator_id}/log")
def api_indicators_add_log(indicator_id: int, body: LoopLogEntryCreate):
    try:
        result = indicators.add_log_entry(
            indicator_id,
            weakness_identified=body.weakness_identified,
            action_taken=body.action_taken,
            entry_status=body.entry_status,
            entry_date=body.entry_date,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if result is None:
        raise HTTPException(status_code=404, detail="Indicator not found")
    return result


# ---------------------------------------------------------------------------
# Accreditation — Standard 2 Curriculum Mapping
# ---------------------------------------------------------------------------
class IloCreate(BaseModel):
    ilo_text: str
    ilo_code: str | None = None


class IloUpdate(BaseModel):
    ilo_text: str | None = None
    ilo_code: str | None = None


class CourseCreate(BaseModel):
    course_name: str


class CoursesImportBulk(BaseModel):
    course_names: list[str]


class MappingSet(BaseModel):
    course_id: int
    ilo_id: int
    mapped: bool


@app.get("/api/curriculum/ilos")
def api_curriculum_list_ilos():
    return curriculum_mapping.list_ilos()


@app.post("/api/curriculum/ilos")
def api_curriculum_create_ilo(body: IloCreate):
    try:
        return curriculum_mapping.create_ilo(body.ilo_text, body.ilo_code)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.patch("/api/curriculum/ilos/{ilo_id}")
def api_curriculum_update_ilo(ilo_id: int, body: IloUpdate):
    try:
        result = curriculum_mapping.update_ilo(ilo_id, ilo_text=body.ilo_text, ilo_code=body.ilo_code)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if result is None:
        raise HTTPException(status_code=404, detail="ILO not found")
    return result


@app.delete("/api/curriculum/ilos/{ilo_id}")
def api_curriculum_delete_ilo(ilo_id: int):
    if not curriculum_mapping.delete_ilo(ilo_id):
        raise HTTPException(status_code=404, detail="ILO not found")
    return {"deleted": True}


@app.get("/api/curriculum/courses")
def api_curriculum_list_courses():
    return curriculum_mapping.list_courses()


@app.post("/api/curriculum/courses")
def api_curriculum_create_course(body: CourseCreate):
    try:
        return curriculum_mapping.create_course(body.course_name)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.delete("/api/curriculum/courses/{course_id}")
def api_curriculum_delete_course(course_id: int):
    if not curriculum_mapping.delete_course(course_id):
        raise HTTPException(status_code=404, detail="Course not found")
    return {"deleted": True}


@app.post("/api/curriculum/courses/import")
def api_curriculum_import_courses(body: CoursesImportBulk):
    return curriculum_mapping.import_courses_bulk(body.course_names, source="import")


@app.post("/api/curriculum/courses/import-excel")
async def api_curriculum_import_courses_excel(file: UploadFile = File(...)):
    """Extract a clean, de-duplicated course list from an uploaded grades workbook."""
    _validate_file(file)
    contents = await _read_limited(file)
    try:
        names = curriculum_mapping.extract_course_names_from_excel(contents)
    except Exception as exc:
        logger.exception("Course extraction failed for curriculum import")
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc
    if not names:
        raise HTTPException(status_code=422, detail="No course names could be extracted from this file")
    return curriculum_mapping.import_courses_bulk(names, source="import-excel")


@app.get("/api/curriculum/matrix")
def api_curriculum_matrix():
    return curriculum_mapping.get_matrix()


@app.post("/api/curriculum/matrix")
def api_curriculum_set_mapping(body: MappingSet):
    curriculum_mapping.set_mapping(body.course_id, body.ilo_id, body.mapped)
    return {"ok": True}


@app.get("/api/curriculum/summary")
def api_curriculum_summary():
    return curriculum_mapping.compute_coverage_summary()


@app.get("/export-curriculum-map-docx")
def export_curriculum_map_docx():
    try:
        from curriculum_map_report import build_curriculum_map_docx
        data = curriculum_mapping.get_export_data()
        docx_bytes = build_curriculum_map_docx(data)
    except Exception as exc:
        logger.exception("Curriculum map DOCX generation failed")
        raise HTTPException(status_code=500, detail="Report generation failed; check server logs") from exc

    return StreamingResponse(
        io.BytesIO(docx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": 'attachment; filename="Curriculum_Map.docx"'},
    )


# ---------------------------------------------------------------------------
# Accreditation — Standard 1 Governance
# ---------------------------------------------------------------------------
class MissionVersionCreate(BaseModel):
    mission_text: str


class StakeholderEntryCreate(BaseModel):
    stakeholder_name: str
    consulted_on: str
    topic: str
    stakeholder_role: str | None = None
    notes: str | None = None


@app.get("/api/governance/mission")
def api_governance_list_mission():
    return governance.list_mission_versions()


@app.get("/api/governance/mission/current")
def api_governance_current_mission():
    current = governance.get_current_mission()
    if current is None:
        raise HTTPException(status_code=404, detail="No mission text has been saved yet")
    return current


@app.post("/api/governance/mission")
def api_governance_create_mission(body: MissionVersionCreate):
    try:
        return governance.create_mission_version(body.mission_text)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/governance/documents")
def api_governance_list_documents(committee_name: str | None = None):
    return governance.list_documents(committee_name=committee_name)


@app.post("/api/governance/documents")
async def api_governance_upload_document(
    file: UploadFile = File(...),
    title: str = Form(...),
    committee_name: str | None = Form(None),
    document_date: str | None = Form(None),
):
    contents = await file.read()
    try:
        return governance.create_document(
            title=title,
            file_bytes=contents,
            original_filename=file.filename or "document",
            storage_dir=_GOVERNANCE_DOCS_DIR,
            committee_name=committee_name,
            document_date=document_date,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@app.get("/api/governance/documents/{doc_id}/file")
def api_governance_get_document_file(doc_id: int):
    doc = governance.get_document(doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
    path = os.path.join(_GOVERNANCE_DOCS_DIR, doc["stored_filename"])
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="Document file missing on server")
    return FileResponse(path, filename=doc["original_filename"])


@app.delete("/api/governance/documents/{doc_id}")
def api_governance_delete_document(doc_id: int):
    if not governance.delete_document(doc_id, _GOVERNANCE_DOCS_DIR):
        raise HTTPException(status_code=404, detail="Document not found")
    return {"deleted": True}


@app.get("/api/governance/stakeholder-log")
def api_governance_list_stakeholder_log():
    return governance.list_stakeholder_log()


@app.post("/api/governance/stakeholder-log")
def api_governance_add_stakeholder_entry(body: StakeholderEntryCreate):
    try:
        return governance.add_stakeholder_entry(
            stakeholder_name=body.stakeholder_name,
            consulted_on=body.consulted_on,
            topic=body.topic,
            stakeholder_role=body.stakeholder_role,
            notes=body.notes,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


# ---------------------------------------------------------------------------
# Chatbot
# ---------------------------------------------------------------------------
class ChatRequest(BaseModel):
    question: str
    context: dict | None = None
    history: list | None = None


@app.post("/api/chat")
async def api_chat(req: ChatRequest):
    """Answer a question about the current analysis result."""
    if not req.question or not req.question.strip():
        return {"answer": "Please type a question."}
    answer = answer_question(req.question.strip(), req.context or {}, req.history or [])
    return {"answer": answer}


# ---------------------------------------------------------------------------
# Dev runner
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=_server.port, reload=True)
